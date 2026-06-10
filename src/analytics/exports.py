"""
Generación de los CSVs y Excel de análisis derivados del master.
"""

import logging
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

BUCKET_LABELS = ["0-1Y", "1-3Y", "3-5Y", "5-10Y", "10Y+"]
BUCKET_BINS   = [float("-inf"), 1, 3, 5, 10, float("inf")]


def _save(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    logger.info(f"Exportado: {path}  ({len(df)} filas)")


def export_top_liquid(df: pd.DataFrame, path: Path, top_n: int = 50) -> None:
    """Top ONs por liquidity_score (solo con monto operado > 0)."""
    cols = [c for c in [
        "ticker", "moneda", "tipo_liquidacion",
        "ultimo_precio", "monto_operado", "volumen_nominal",
        "cantidad_operaciones", "spread_abs", "spread_pct",
        "years_to_maturity", "liquidity_score",
    ] if c in df.columns]

    result = (
        df[cols]
        .query("monto_operado > 0")
        .sort_values("liquidity_score", ascending=False)
        .head(top_n)
        .reset_index(drop=True)
    )
    _save(result, path)


def export_top_spreads(df: pd.DataFrame, path: Path) -> None:
    """
    ONs con bid y ask presentes, ordenadas de menor a mayor spread_pct.
    Incluye columna 'spread_cuartil' para identificar rápidamente la zona.
    """
    cols = [c for c in [
        "ticker", "moneda", "tipo_liquidacion",
        "precio_compra", "precio_venta",
        "spread_abs", "spread_pct",
        "monto_operado", "years_to_maturity", "liquidity_score",
    ] if c in df.columns]

    result = (
        df[cols]
        .dropna(subset=["spread_abs"])
        .sort_values("spread_pct")
        .reset_index(drop=True)
    )

    if not result.empty:
        result["spread_cuartil"] = pd.qcut(
            result["spread_pct"],
            q=4,
            labels=["Q1_tight", "Q2", "Q3", "Q4_wide"],
            duplicates="drop",
        )

    _save(result, path)


def export_maturity_buckets(df: pd.DataFrame, path: Path) -> None:
    """
    Estadísticas agregadas por bucket de vencimiento.
    Incluye todos los registros (con y sin trade).
    """
    if "years_to_maturity" not in df.columns:
        logger.warning("years_to_maturity no disponible — omitiendo maturity_buckets")
        return

    df = df.copy()
    df["bucket"] = pd.cut(
        df["years_to_maturity"],
        bins=BUCKET_BINS,
        labels=BUCKET_LABELS,
        right=False,
    )

    agg = (
        df.groupby("bucket", observed=True)
        .agg(
            cantidad=("ticker", "count"),
            con_trade=("ultimo_precio", lambda s: (s > 0).sum()),
            con_bid_ask=("spread_abs", lambda s: s.notna().sum()),
            monto_operado_total=("monto_operado", "sum"),
            monto_operado_avg=("monto_operado", "mean"),
            spread_pct_avg=("spread_pct", "mean"),
            spread_pct_median=("spread_pct", "median"),
            liquidity_score_avg=("liquidity_score", "mean"),
            years_to_maturity_avg=("years_to_maturity", "mean"),
        )
        .round(4)
        .reset_index()
    )

    _save(agg, path)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _sheet_todas(df: pd.DataFrame) -> pd.DataFrame:
    cols = [c for c in [
        "ticker", "moneda", "tipo_liquidacion",
        "ultimo_precio", "precio_compra", "precio_venta",
        "spread_abs", "spread_pct",
        "monto_operado", "volumen_nominal", "cantidad_operaciones",
        "years_to_maturity", "fecha_vencimiento", "liquidity_score",
    ] if c in df.columns]
    return (
        df[cols]
        .sort_values(["monto_operado", "spread_pct", "years_to_maturity"],
                     ascending=[False, True, True], na_position="last")
        .reset_index(drop=True)
    )


def _sheet_top_liquidez(df: pd.DataFrame, top_n: int = 50) -> pd.DataFrame:
    cols = [c for c in [
        "ticker", "moneda", "tipo_liquidacion",
        "ultimo_precio", "monto_operado", "volumen_nominal",
        "cantidad_operaciones", "spread_abs", "spread_pct",
        "years_to_maturity", "liquidity_score",
    ] if c in df.columns]
    return (
        df[cols]
        .query("monto_operado > 0")
        .sort_values("liquidity_score", ascending=False)
        .head(top_n)
        .reset_index(drop=True)
    )


def _sheet_menor_spread(df: pd.DataFrame, top_n: int = 50) -> pd.DataFrame:
    cols = [c for c in [
        "ticker", "moneda", "tipo_liquidacion",
        "precio_compra", "precio_venta",
        "spread_abs", "spread_pct",
        "monto_operado", "years_to_maturity", "liquidity_score",
    ] if c in df.columns]
    return (
        df[cols]
        .dropna(subset=["spread_abs"])
        .sort_values("spread_pct")
        .head(top_n)
        .reset_index(drop=True)
    )


def _sheet_mayor_spread(df: pd.DataFrame, top_n: int = 50) -> pd.DataFrame:
    cols = [c for c in [
        "ticker", "moneda", "tipo_liquidacion",
        "precio_compra", "precio_venta",
        "spread_abs", "spread_pct",
        "monto_operado", "years_to_maturity", "liquidity_score",
    ] if c in df.columns]
    return (
        df[cols]
        .dropna(subset=["spread_abs"])
        .sort_values("spread_pct", ascending=False)
        .head(top_n)
        .reset_index(drop=True)
    )


def _sheet_por_moneda(df: pd.DataFrame) -> pd.DataFrame:
    agg = (
        df.groupby("moneda", observed=True)
        .agg(
            cantidad=("ticker", "count"),
            con_trade=("ultimo_precio", lambda s: (s > 0).sum()),
            con_bid_ask=("spread_abs", lambda s: s.notna().sum()),
            monto_operado_total=("monto_operado", "sum"),
            monto_operado_avg=("monto_operado", "mean"),
            spread_pct_avg=("spread_pct", "mean"),
            spread_pct_median=("spread_pct", "median"),
            liquidity_score_avg=("liquidity_score", "mean"),
            years_to_maturity_avg=("years_to_maturity", "mean"),
        )
        .round(4)
        .reset_index()
        .sort_values("monto_operado_total", ascending=False)
    )
    return agg


def _style_sheet(ws, header_color: str = "1F4E79") -> None:
    """Aplica formato al encabezado y ajusta ancho de columnas."""
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", fgColor=header_color)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    # Auto-ancho basado en el contenido (máx 30)
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

    ws.freeze_panes = "A2"


SHEETS = [
    ("Todas las ONs",  _sheet_todas),
    ("Top liquidez",   _sheet_top_liquidez),
    ("Menor spread",   _sheet_menor_spread),
    ("Mayor spread",   _sheet_mayor_spread),
    ("Por moneda",     _sheet_por_moneda),
]


def export_excel(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, build_fn in SHEETS:
            sheet_df = build_fn(df)
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            _style_sheet(writer.sheets[sheet_name])

    logger.info(f"Exportado: {path}  ({len(SHEETS)} hojas)")


# ---------------------------------------------------------------------------
# Issuer enrichment exports
# ---------------------------------------------------------------------------

def export_ons_enriched(df: pd.DataFrame, path: Path) -> None:
    """
    Vista limpia para análisis externo: un registro por ON activa.
    Columnas pedidas: symbol, issuer, cuit, currency, price, volume,
                      spread_pct, years_to_maturity + extras útiles.
    """
    cols_map = {
        "ticker":            "symbol",
        "issuer":            "issuer",
        "cuit":              "cuit",
        "sector":            "sector",
        "moneda":            "currency",
        "tipo_liquidacion":  "settlement",
        "ultimo_precio":     "price",
        "precio_compra":     "bid",
        "precio_venta":      "ask",
        "monto_operado":     "volume",
        "spread_abs":        "spread_abs",
        "spread_pct":        "spread_pct",
        "years_to_maturity": "years_to_maturity",
        "fecha_vencimiento": "maturity_date",
        "liquidity_score":   "liquidity_score",
        "is_mapped":         "is_mapped",
    }
    present = {k: v for k, v in cols_map.items() if k in df.columns}
    result = (
        df[list(present.keys())]
        .rename(columns=present)
        .sort_values(["volume", "spread_pct"], ascending=[False, True], na_position="last")
        .reset_index(drop=True)
    )
    _save(result, path)


def export_top_issuers(df_issuers: pd.DataFrame, path: Path) -> None:
    """Top emisores por volumen total."""
    _save(df_issuers, path)
